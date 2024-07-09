import os
import google.oauth2.credentials
import google_auth_oauthlib.flow
import googleapiclient.discovery
import googleapiclient.errors
from googleapiclient.http import MediaFileUpload
from datetime import datetime, timedelta
import time
import openpyxl


# Set the scopes and API service name/version
SCOPES = ["https://www.googleapis.com/auth/youtube.upload"]
API_SERVICE_NAME = "youtube"
API_VERSION = "v3"


# Authenticate and authorize
def get_authenticated_service():
    flow = google_auth_oauthlib.flow.InstalledAppFlow.from_client_secrets_file(
        "client_secret_83190869396-krraupejd6asno6t6ctok8qtf82st9s5.apps.googleusercontent.com.json", SCOPES)
    credentials = flow.run_local_server(port=0)
    return googleapiclient.discovery.build(API_SERVICE_NAME, API_VERSION, credentials=credentials)

def upload_video(video_file, title, description, tags, publish_time):
    body = {
        "snippet": {
            "title": title,
            "description": description,
            "tags": tags,
            "categoryId": "27"  # Category ID for 'People & Blogs'
        },
        "status": {
            "privacyStatus": "private",
            "publishAt": publish_time  # Properly formatted publish time
        }
    }

    media = MediaFileUpload(video_file, chunksize=-1, resumable=True)
    request = youtube.videos().insert(
        part="snippet,status",
        body=body,
        media_body=media
    )

    response = request.execute()
    print(f"Uploaded video with ID: {response['id']}")
if __name__ == "__main__":
    
    start_number = 1
    no_vids=256

    path = "C:\\Users\\sudha\\Documents\\GitHub\\video-poster\\ai_shorts_psychology_with_separate_tags.xlsx"
    wb = openpyxl.load_workbook(path)
    sheet = wb.active

    date_modifier = 10 #put starting day
    month_modifier = 7 #put starting month
    youtube = get_authenticated_service()
    for i in range(start_number,start_number+no_vids):
        I=str(i)
        video_file = "C:\\Users\\sudha\\Documents\\GitHub\\video-poster\\videos to post\\"+I+".mp4"
        titleTemp = sheet.cell(row = i+1,column = 1)
        title = str(titleTemp.value)

        descriptionTemp = sheet.cell(row = i+1,column = 2)
        description = str(descriptionTemp.value)

        tags = []
        for j in range(1,16):
            appendable  = sheet.cell(row = i+1, column = j+2)
            tags.append(str(appendable.value))

        
        remainder = i % 3

        if (remainder == 0):
            publish_time = datetime(2024, month_modifier, date_modifier, 4, 0, 0).isoformat() + "Z"
            sleep_amount = 10
        elif(remainder == 1):
            publish_time = datetime(2024, month_modifier, date_modifier, 12, 0, 0).isoformat() + "Z"
            sleep_amount = 10
        elif(remainder == 2):
            publish_time = datetime(2024, month_modifier, date_modifier, 20, 0, 0).isoformat() + "Z"
            date_modifier += 1
            sleep_amount = 60
            if (date_modifier >=30):
                date_modifier = 1
                month_modifier += 1
        else:
            print("error")

        print(video_file, title, description, tags, publish_time)

        upload_video(video_file, title, description, tags, publish_time)
        time.sleep(sleep_amount)