import os
import google.oauth2.credentials
import google_auth_oauthlib.flow
import googleapiclient.discovery
import googleapiclient.errors
from googleapiclient.http import MediaFileUpload
from datetime import datetime, timedelta
import time
import openpyxl

# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Set the scopes and API service name/version
SCOPES = ["https://www.googleapis.com/auth/youtube.upload"]
API_SERVICE_NAME = "youtube"
API_VERSION = "v3"


# Authenticate and authorize
def get_authenticated_service(Input):
    
    if (Input ==1):
        flow = google_auth_oauthlib.flow.InstalledAppFlow.from_client_secrets_file(
        "client_secret_83190869396-krraupejd6asno6t6ctok8qtf82st9s5.apps.googleusercontent.com.json", SCOPES)
        credentials = flow.run_local_server(port=0)
    elif (Input == 2):
        flow = google_auth_oauthlib.flow.InstalledAppFlow.from_client_secrets_file(
        "client_secret_949978928070-pu7g01jbd3ndco42rdvqjjv3ch2nhsnc.apps.googleusercontent.com.json", SCOPES)
        credentials = flow.run_local_server(port=0)
    elif(Input == 3):
        flow = google_auth_oauthlib.flow.InstalledAppFlow.from_client_secrets_file(
        "client_secret_967439837831-26u1s4qfvlrurioka0ltbkd6kkhsl6gn.apps.googleusercontent.com.json", SCOPES)
        credentials = flow.run_local_server(port=0)
    else:
        print("Finished")
        
    return googleapiclient.discovery.build(API_SERVICE_NAME, API_VERSION, credentials=credentials)

    
def upload_video(video_file, title, description, tags, publish_time):
    body = {
        "snippet": {
            "title": title,
            "description": description,
            "tags": tags,
            "categoryId": "27"  # Category ID for 'People & Blogs'
        },
        "status": {
            "privacyStatus": "private",
            "publishAt": publish_time  # Properly formatted publish time
        }
    }

    media = MediaFileUpload(video_file, chunksize=-1, resumable=True)
    request = youtube.videos().insert(
        part="snippet,status",
        body=body,
        media_body=media
    )

    response = request.execute()
    print(f"Uploaded video with ID: {response['id']}")
if __name__ == "__main__":
    
    
    
    date_modifier = 4 #put starting day
    month_modifier = 8 #put starting month
    

    no_vids=256

    fileName = "ai_shorts_psychology_with_separate_tags.xlsx"
    path = os.path.join(script_dir, fileName)
    wb = openpyxl.load_workbook(path)
    sheet1 = wb["Sheet1"]
    sheet2 = wb["Sheet2"]


    start_number = int(sheet2.cell(row = 1, column = 1).value)
    
    appNo = 3
    count = 1
    
    youtube = get_authenticated_service(appNo)
    
    for j in range(1, start_number+1):
        re = j%3
        if re == 2:
            date_modifier += 1
            if (date_modifier >=30):
                date_modifier = 1
                month_modifier += 1

    for i in range(start_number,start_number+no_vids):
        I=str(i)
        videoName = "videos to post\\"+I+".mp4"
        video_file = os.path.join(script_dir, videoName)
        titleTemp = sheet1.cell(row = i+1,column = 1)
        title = str(titleTemp.value)

        descriptionTemp = sheet1.cell(row = i+1,column = 2)
        description = str(descriptionTemp.value)

        tags = []
        for j in range(1,16):
            appendable  = sheet1.cell(row = i+1, column = j+2)
            tags.append(str(appendable.value))

        
        remainder = i % 3

        if (remainder == 0):
            publish_time = datetime(2024, month_modifier, date_modifier, 4, 0, 0).isoformat() + "Z"
        elif(remainder == 1):
            publish_time = datetime(2024, month_modifier, date_modifier, 12, 0, 0).isoformat() + "Z"
        elif(remainder == 2):
            publish_time = datetime(2024, month_modifier, date_modifier, 20, 0, 0).isoformat() + "Z"
            date_modifier += 1
            if (date_modifier >=30):
                date_modifier = 1
                month_modifier += 1
        else:
            print("error")


        upload_video(video_file, title, description, tags, publish_time)
        print(video_file, title, description, tags, publish_time)
        
        
        
        
        if (count ==7):
            count = 0
            appNo+=1
            youtube = get_authenticated_service(appNo)
        else:
            count +=1
        sheet2['A1'] = i
        wb.save(path)
        
        time.sleep(5)